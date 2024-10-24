##
# Date of change: July 10th, 2022

## -----------------------------------------------------------------------------------------------##
#                                                                                                  #
## ---------------------------------------------------------------------------------------------- ##


# basics
import numpy as np                      # diff, mean, nanmean, sqrt, var    # float64, int64
from openpyxl import Workbook
from pandas import read_csv
from scipy.stats import skewtest

# my modules
from saodm_convert_excel import csv_path    # path where the converted files shall be stored
findings_path = 'FINDINGS/'        # path to store the results (csv, xlsx)
from saodm_useful import calc_fano_factor, read_csv_columns
#from saodm_useful import calc_spikelet_regularity_and_length, calc_rr_sd1_sd2


####################################################################################################


##
# @param case_id: 
# @param C_SIF_s: 
# @param C_CHEM_s: 
# @param SECTION_SIZE_IN_S: 
def analyze_and_compare(case_id: str, C_SIF_start: float, C_CHEM_start: float,
                        SECTION_SIZE_IN_S: float):
    '''
    '''
    file_spikes_SIF = csv_path + case_id + '_spikes_SIF.csv'
    file_spikes_CHEM = csv_path + case_id + '_spikes_CHEM.csv'
    file_isi_SIF = csv_path + case_id + '_isi_SIF.csv'
    file_isi_CHEM = csv_path + case_id + '_isi_CHEM.csv'

    try:
        spikes_SIF = read_csv_columns(file_spikes_SIF, ['Time'])['Time']
        spikes_CHEM = read_csv_columns(file_spikes_CHEM, ['Time'])['Time']
        isi_SIF = read_csv_columns(file_isi_SIF, ['Time'])['Time']
        isi_CHEM = read_csv_columns(file_isi_CHEM, ['Time'])['Time']
    except Exception as e:
        print(type(e))
        print(e)
    finally:
        del file_spikes_SIF
        del file_spikes_CHEM
        del file_isi_SIF
        del file_isi_CHEM

    ############################################################################

    dict_features = dict()

    i_numSpikesSIF, i_numSpikesCHEM = len(spikes_SIF), len(spikes_CHEM)
    s_numSpikesSIF, s_numSpikesCHEM = '%i' % i_numSpikesSIF, '%i' % i_numSpikesCHEM
    dict_features['Number of Spikes [#]'] = (s_numSpikesSIF, s_numSpikesCHEM)
    s_480SIF = '%.3f' % (SECTION_SIZE_IN_S/(i_numSpikesSIF-1)) if i_numSpikesSIF > 1 else ""    # empty
    s_480CHEM = '%.3f' % (SECTION_SIZE_IN_S/(i_numSpikesCHEM-1)) if i_numSpikesCHEM > 1 else "" # empty
    dict_features['%i by (Number of Spikes - 1) [s]' % SECTION_SIZE_IN_S] = (s_480SIF, s_480CHEM)
    del i_numSpikesSIF, i_numSpikesCHEM
    del s_numSpikesSIF, s_numSpikesCHEM
    del s_480SIF, s_480CHEM

    s_meanSIF, s_meanCHEM = '%.3f' % np.mean(isi_SIF), '%.3f' % np.mean(isi_CHEM)
    s_medianSIF, s_medianCHEM = '%.3f' % np.median(isi_SIF), '%.3f' % np.median(isi_CHEM)
    s_skewtestSIF = '%.3f' % skewtest(isi_SIF)[1] if len(isi_SIF) > 7 else ""
    s_skewtestCHEM = '%.3f' % skewtest(isi_CHEM)[1] if len(isi_CHEM) > 7 else ""
    dict_features['ISI Mean [s]'] = (s_meanSIF, s_meanCHEM)
    dict_features['ISI Median [s]'] = (s_medianSIF, s_medianCHEM)
    dict_features['ISI Skewtest []'] = (s_skewtestSIF, s_skewtestCHEM)
    del s_meanSIF, s_meanCHEM
    del s_medianSIF, s_medianCHEM

    f_isiMinSIF, f_isiMinCHEM = min(isi_SIF), min(isi_CHEM)
    s_isiMinSIF, s_isiMinCHEM = '%.3f' % f_isiMinSIF, '%.3f' % f_isiMinCHEM
    dict_features['ISI Minimum [s]'] = (s_isiMinSIF, s_isiMinCHEM)
    s_freqMaximumSIF, s_freqMaximumCHEM = '%.3f' % (1./f_isiMinSIF), '%.3f' % (1./f_isiMinCHEM)
    dict_features['Frequency Maximum [Hz]'] = (s_freqMaximumSIF, s_freqMaximumCHEM)
    f_isiMaxSIF, f_isiMaxCHEM = max(isi_SIF), max(isi_CHEM)
    s_isiMaxSIF, s_isiMaxCHEM = '%.3f' % f_isiMaxSIF, '%.3f' % f_isiMaxCHEM
    dict_features['ISI Maximum [s]'] = (s_isiMaxSIF, s_isiMaxCHEM)
    s_diffMaxMinSIF = '%.3f' % (f_isiMaxSIF-f_isiMinSIF)
    s_diffMaxMinCHEM = '%.3f' % (f_isiMaxCHEM-f_isiMinCHEM)
    dict_features['ISI Difference [s]'] = (s_diffMaxMinSIF, s_diffMaxMinCHEM)
    del f_isiMinSIF, f_isiMinCHEM
    del s_isiMinSIF, s_isiMinCHEM
    del s_freqMaximumSIF, s_freqMaximumCHEM
    del f_isiMaxSIF, f_isiMaxCHEM
    del s_isiMaxSIF, s_isiMaxCHEM
    del s_diffMaxMinSIF, s_diffMaxMinCHEM

    ## HRV metrics #############################################################

    # SDNN: Standard Deviation of the NN intervals, i.e. square root of the variance.
    # Statistically not significant, when recording duration differs [Circulation 1996, Vol. 93]
    f_varSIF, f_varCHEM = np.var(isi_SIF), np.var(isi_CHEM)
    s_varSIF, s_varCHEM = '%.3f' % f_varSIF, '%.3f' % f_varCHEM
    dict_features['ISI Variance [s²]'] = (s_varSIF, s_varCHEM)
    s_stdevSIF, s_stdevCHEM = '%.3f' % np.sqrt(f_varSIF), '%.3f' % np.sqrt(f_varCHEM)
    dict_features['ISI Standard Deviation [s]'] = (s_stdevSIF, s_stdevCHEM)
    del f_varSIF, f_varCHEM
    del s_varSIF, s_varCHEM
    del s_stdevSIF, s_stdevCHEM

    # RMSSD: Root Mean Square of Successive Differences.
    # most commonly used measure, good statistical properties
    def rmssd(isi_list):
        n = len(isi_list)
        if n > 1:
            squaredSum = 0.
            for i in range(n-1):
                squaredSum += ((isi_list[i+1] - isi_list[i]) ** 2)
            return '%.3f' % np.sqrt(1/(n-1) * squaredSum)
        return ""   # empty

    rmssdSIF, rmssdCHEM = rmssd(isi_SIF), rmssd(isi_CHEM)
    dict_features['RMSSD [s]'] = (rmssdSIF, rmssdCHEM)
    del rmssdSIF, rmssdCHEM

    # Fano factor
    fano48sSIF = '%.3f' % calc_fano_factor(spikes_SIF, C_SIF_start, 48.)
    fano48sCHEM = '%.3f' % calc_fano_factor(spikes_CHEM, C_CHEM_start, 48.)
    dict_features['Fano factor t=48s []'] = (fano48sSIF, fano48sCHEM)
    del fano48sSIF, fano48sCHEM

    # Poincaré
    """
    (_, f_sd1, f_sd2) = calc_rr_sd1_sd2(isi_SIF)
    sd1SIF = '%.3f' % f_sd1
    sd2SIF = '%.3f' % f_sd2
    sd1bysd2SIF = '%.3f' % (f_sd1/f_sd2) if f_sd2 != 0 else ""  # empty
    (_, f_sd1, f_sd2) = calc_rr_sd1_sd2(isi_CHEM)
    sd1CHEM = '%.3f' % f_sd1
    sd2CHEM = '%.3f' % f_sd2
    sd1bysd2CHEM = '%.3f' % (f_sd1/f_sd2) if f_sd2 != 0 else "" # empty
    dict_features['SD1'] = (sd1SIF, sd1CHEM)
    dict_features['SD2'] = (sd2SIF, sd2CHEM)
    dict_features['SD1 by SD2'] = (sd1bysd2SIF, sd1bysd2CHEM)
    del sd1SIF, sd2SIF, sd1bysd2SIF
    del sd1CHEM, sd2CHEM, sd1bysd2CHEM
    """

    # Spikelets
    """    
    (list_regSIF, list_lenSIF) = calc_spikelet_regularity_and_length(isi_SIF)
    reg_meanSIF = '%.3f' % np.mean(list_regSIF)              # mu
    len_meanSIF = '%.3f' % np.mean(list_lenSIF)              # mu
    reg_medianSIF = '%.3f' % np.median(list_regSIF)          # p50
    len_medianSIF = '%.3f' % np.median(list_lenSIF)          # p50
    reg_stdevSIF = '%.3f' % np.sqrt(np.var(list_regSIF))     # sigma
    len_stdevSIF = '%.3f' % np.sqrt(np.var(list_lenSIF))     # sigma
    (list_regCHEM, list_lenCHEM) = calc_spikelet_regularity_and_length(isi_CHEM)
    reg_meanCHEM = '%.3f' % np.mean(list_regCHEM)            # mu
    len_meanCHEM = '%.3f' % np.mean(list_lenCHEM)            # mu
    reg_medianCHEM = '%.3f' % np.median(list_regCHEM)        # p50
    len_medianCHEM = '%.3f' % np.median(list_lenCHEM)        # p50
    reg_stdevCHEM = '%.3f' % np.sqrt(np.var(list_regCHEM))   # sigma
    len_stdevCHEM = '%.3f' % np.sqrt(np.var(list_lenCHEM))   # sigma
    dict_features['S.let Regul. Mean'] = (reg_meanSIF, reg_meanCHEM)
    dict_features['S.let Regul. Median'] = (reg_medianSIF, reg_medianCHEM)
    dict_features['S.let Regul. STDEV'] = (reg_stdevSIF, reg_stdevCHEM)
    dict_features['S.let Length Mean'] = (len_meanSIF, len_meanCHEM)
    dict_features['S.let Length Median'] = (len_medianSIF, len_medianCHEM)
    dict_features['S.let Length STDEV'] = (len_stdevSIF, len_stdevCHEM)
    del list_regSIF, list_lenSIF
    del list_regCHEM, list_lenCHEM
    del reg_meanSIF, reg_medianSIF, reg_stdevSIF
    del len_meanSIF, len_medianSIF, len_stdevSIF
    del reg_meanCHEM, reg_medianCHEM, reg_stdevCHEM
    del len_meanCHEM, len_medianCHEM, len_stdevCHEM
    """

    ####################################

    ## add new metrics here


    ####################################


    # store features for case_id in a single csv file
    with open(findings_path + case_id + '.csv', 'wt', encoding='utf8') as fw:
        string = ''
        for tag in dict_features.keys():
            string += tag + ';'
        string = string[:-1] + '\n'     # delete last ; and replace it by \n
        fw.write(string)
        string = ''
        for tag in dict_features.keys():
            string += dict_features[tag][0] + ';'
        string = string[:-1] + '\n'
        fw.write(string)
        string = ''
        for tag in dict_features.keys():
            string += dict_features[tag][1] + ';'
        string = string[:-1] + '\n'
        fw.write(string)

    del spikes_SIF
    del spikes_CHEM
    del isi_SIF
    del isi_CHEM
    del C_SIF_start
    del C_CHEM_start
    del dict_features


##
# @param dict_grouped: 
def merge_findings_table(dict_grouped):
    '''
    Merge cases to xlsx table.
    '''
    wb = Workbook()
    ws = wb.active
    ws.title = "Findings"

    list_average = list()
    next_row = 1
    for tupl in sorted(dict_grouped.keys()):
        # Preparation ######################################
        columns = read_csv(findings_path + dict_grouped[tupl][0] + '.csv',
            delimiter=';')
        list_features = list(dict(columns).keys())
        list_average = [0] * 2*len(list_features)
        for i in range(2*len(list_features)):
            list_average[i] = list()

        # Line 1 ###########################################
        # e.g. "Control GO", "Diabetic MG"
        next_row += 1
        ws.cell(row=next_row, column=2, value=tupl[0] + ' ' + tupl[1])
        next_row += 1

        # Line 2 ###########################################
        # i.e. "condition; Number of Spikes; ..."
        ws.cell(row=next_row, column=3, value='feature')
        next_col = 4
        for i in range(len(list_features)):
            ws.cell(row=next_row, column=next_col, value=list(list_features)[i])
            next_col += 2

        # Line 2 ###########################################
        # e.g. "C_GO_04" followed by tuples numbers
        next_row += 1
        ws.cell(row=next_row, column=3, value='condition')
        next_col = 4
        for i in range(len(list_features)):
            ws.cell(row=next_row, column=next_col, value='SIF')
            ws.cell(row=next_row, column=next_col+1, value=tupl[1])
            next_col += 2

        # Line 3 to n ######################################
        for case_id in dict_grouped[tupl]:
            next_row += 1
            ws.cell(row=next_row, column=3, value=case_id)
            columns = read_csv(findings_path + case_id + '.csv', delimiter=';')
            next_col = 4
            for i, feature_id in enumerate(list_features):
                ws.cell(row=next_row, column=next_col, value=columns[feature_id][0])
                ws.cell(row=next_row, column=next_col+1, value=columns[feature_id][1])
                list_average[next_col-4].append(columns[feature_id][0])
                list_average[next_col-4+1].append(columns[feature_id][1])
                next_col += 2

        next_row += 1
        ws.cell(row=next_row, column=3, value='average')
        next_col = 4
        for i in range(len(list_average)):
            arr = list_average[i]
            # nanmean excludes nan (not a number) and calculated the mean only for valid numbers
            ws.cell(row=next_row, column=next_col, value=round(np.nanmean(arr), 4))
            #if isinstance(arr[0], np.int64) or isinstance(arr[0], np.float64):
            #    ws.cell(row=next_row, column=next_col, value=round(np.nanmean(arr), 4))
            #else:
            #    ws.cell(row=next_row, column=next_col, value='---')
            next_col += 1

        next_row += 1

    wb.save(findings_path + 'findings.xlsx')
    if wb is not None:
        wb.close()

