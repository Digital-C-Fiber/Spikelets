##
# Date of change: July 10th, 2022

## -----------------------------------------------------------------------------------------------##
# This file is used as control module to load data from well known excel files, convert to CSV and #
# leave the rest to saodm_main.                                                                    #
# The Time shall be consistant for all 2 to 3 files "spikes_all", "comment" and "temperature".     #
## -----------------------------------------------------------------------------------------------##
# This package heavily depends on a somewhat reasonable structure of the xlsx files, thou a lot of #
# "variance" has to be counteracted, like in which column the data is in. Other structure is       #
# assumed to be present in each xlsx file like "Time, Text, Freq (Hz), ID" in row 4 and a Comment  #
# Table with "heat, cold, SIF 10min, 8' SIF, ***, SIF kurz, heat, cold" and so on to the "Text"    #
# Coloum.                                                                                          #
## ---------------------------------------------------------------------------------------------- ##

import numpy as np                      # diff
from openpyxl.cell import Cell
import openpyxl.utils.exceptions

# my modules
excel_path = 'EXCEL/'              # path where the test cases (excel) are stored
csv_path = 'CSV/'                  # path where the converted files shall be stored
from saodm_useful import read_csv_columns


####################################################################################################


##
# @param    cell: an excel cell content
# @return   value of cell but ' ' if value is None or 'None'
def _val(cell: Cell):
    '''
    Sometimes the excel cell has no value (None). To avoid these in lists,
    replace them by the other empty symbol ' ' (space).
    '''
    val = cell.value
    return ' ' if ((val is None) or (val == 'None')) else val


##
# @param
# @param
def save_to_file(file: str, Time_list=[]):
    '''
    '''
    with open(file=file, mode='wt') as fw:
        fw.write('Time\n')
        for time in Time_list:
            fw.write('%.4f\n' % (time))
                
                
##
class Sheet:
    '''
    Class to manage single excel file (workbook) with all tables (sheets).
    '''

    ##
    # @param
    def __init__(self, file_name: str):
        self.workbook = None
        #self.sheet = None
        try:
            self.workbook = openpyxl.load_workbook(file_name)
        except FileNotFoundError:
            print('\n    FileNotFoundError:', file_name, '\n')
            exit(0)
        except Exception as e:
            raise e

    def __del__(self):
        self.sheet = None
        if self.workbook is not None:
            self.workbook.close()
            del self.workbook

    def __repr__(self):
        return ''   #'\n%12s: %6.1fs' % (self.name, self.values[-1] - self.values[0])

    def __str__(self):
        return self.__repr__()

    ##
    # @param    sheet_no: [1:n] rows and cols start with "1", so should sheets.
    def open_sheet_no(self, sheet_no: int):
        ''' sheet_no is [1:n] '''
        # sheetnames is [0:n-1], therefore sheet_no-1
        self.sheet = None
        try:
            self.sheet = self.workbook[self.workbook.sheetnames[sheet_no-1]]
        except Exception as e:
            raise e

    ##
    # @param    letter:
    # @param    start_row: first row to return
    # @return
    def get_column(self, letter: str, start_row=1):
        ret = list()
        letter = letter.upper()
        for i, cell in enumerate(self.sheet[letter]):
            if i+1 >= start_row:
                val = _val(cell)
                if val == ' ':
                    break
                ret.append(val)
        return ret

    ##
    # @param
    # @param
    # @param
    def comments_to_csv(self, case_id: str, cmt_Time_list=[], cmt_Text_list=[]):
        '''
        '''
        with open(file=csv_path + case_id + '_comments.csv', mode='wt') as fw:
            fw.write('Time;Text' + '\n')
            for (time, text) in zip(cmt_Time_list, cmt_Text_list):
                fw.write('%.4f;%s\n' % (time, text))

    ##
    # @param
    # @param
    def filter1_to_csv(self, case_id: str, ftr_Time_list=[]):
        '''
        '''
        with open(file=csv_path + case_id + '_filter1.csv', mode='wt') as fw:
            fw.write('Time\n')
            for time in ftr_Time_list:
                fw.write('%.4f\n' % (time))

    ##
    # @param
    # @param
    # @param
    # @param
    # @param
    # @param
    def temperature_to_csv(self, case_id: str, tem_Time_list=[], tem_Value_list=[],
                           tem_hour=0, tem_min=0, tem_sec=0):
        '''
        '''
        # reformat Temperature Time
        date_offset = int(tem_Time_list[0])
        tem_Time_list = [x - date_offset for x in tem_Time_list]    # removes date
        tem_Time_list = [x*24 - tem_hour for x in tem_Time_list]    # removes hour
        tem_Time_list = [x*60 - tem_min for x in tem_Time_list]     # removes minute
        tem_Time_list = [x*60 - tem_sec for x in tem_Time_list]     # removes second

        with open(file=csv_path + case_id + '_temperature.csv', mode='wt') as fw:
            fw.write('Time;Value' + '\n')
            for (time, value) in zip(tem_Time_list, tem_Value_list):
                fw.write('%.4f;%.2f\n' % (time, value))


####################################################################################################


##
# @param
def probe_workbook_and_exit(file_name: str):
    '''
    Opens every table and prints the first 10 lines.
    Afterwards closes program completely to let human modify code.
    '''
    sheet = Sheet(excel_path + file_name)
    print()
    print(sheet.workbook.sheetnames)

    for sheet_name in sheet.workbook.sheetnames:
        print()
        print(sheet_name)
        sheet = sheet.workbook[sheet_name]
        print('Max Row:', sheet.max_row, ', Max Col:', sheet.max_column)

        if True:
            count = 0
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                print(i+1, [' ' if x==None else x for x in row])
                count += 1
                if count > 9:
                    break

    sheet.workbook.close()
    exit(0)


##
# @param
def convert_excel_to_csv(case_id: str, relevant_size: float):
    '''
    Convert content of file hidden behind case_id from excel to csv.
    '''

    sheet = None
    ftr_Time_list = list()
    cmt_Time_list = list()
    cmt_Text_list = list()
    tem_Time_list = list()
    tem_Value_list = list()
    
    
    print('File:', case_id)

    sheet = Sheet(excel_path  + case_id + '.xlsx')
    sheet.open_sheet_no(1)
    ftr_Time_list = sheet.get_column('J', 5)
    cmt_Time_list = sheet.get_column('E', 7)[:7]
    cmt_Text_list = sheet.get_column('F', 7)[:7]
    cmt_Text_list[2] = 'SIF 10min'                          # replace to unify
    cmt_Text_list[3] = "8' SIF"                             # replace to unify
    cmt_Text_list[4] = 'GO 10mM 10min'                      # replace to unify
    cmt_Time_list.insert(5, cmt_Time_list[5] - 480.)        # set new 8' CHEM
    cmt_Text_list.insert(5, "8' CHEM")                      # set new 8' CHEM
    tem_Time_list = sheet.get_column('B', 7)
    tem_Value_list = sheet.get_column('C', 7)
    sheet.comments_to_csv(case_id, cmt_Time_list, cmt_Text_list)
    sheet.filter1_to_csv(case_id, ftr_Time_list)
    #sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 11, 55, 17)
    # correction of wrong temperature times
    sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 11-1, 55-41, 17-50.33)
    del sheet
    sheet = None

    '''
    if case_id == 'C_GO_04':
        print('- Case:', case_id)
        # CONTROL GO 04
        # spikes
        # possibly wrong Time for temperature
        # contains "absaugen" instead of "SIF kurz"
        # missing "stopp"
        # stop with second "heat", missing second "cold" and "v.frey"
        # contains additional entries
        #sheet = Sheet(excel_path + 'CONTROL/' + CON CMH4.xlsx')
        sheet = Sheet(excel_path + 'CONTROL/' + 'C_GO_04.xlsx')
        sheet.open_sheet_no(1)
        ftr_Time_list = sheet.get_column('J', 5)
        cmt_Time_list = sheet.get_column('E', 7)[:7]
        cmt_Text_list = sheet.get_column('F', 7)[:7]
        cmt_Text_list[2] = 'SIF 10min'                          # replace to unify
        cmt_Text_list[3] = "8' SIF"                             # replace to unify
        cmt_Text_list[4] = 'GO 10mM 10min'                      # replace to unify
        cmt_Time_list.insert(5, cmt_Time_list[5] - 480.)        # set new 8' CHEM
        cmt_Text_list.insert(5, "8' CHEM")                      # set new 8' CHEM
        tem_Time_list = sheet.get_column('B', 7)
        tem_Value_list = sheet.get_column('C', 7)
        sheet.comments_to_csv(case_id, cmt_Time_list, cmt_Text_list)
        sheet.filter1_to_csv(case_id, ftr_Time_list)
        #sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 11, 55, 17)
        # correction of wrong temperature times
        sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 11-1, 55-41, 17-50.33)
        del sheet
        sheet = None

    elif case_id == 'C_GO_14':
        print('- Case:', case_id)
        # CONTROL GO 14
        # spikes
        # temperature
        # missing "stopp"
        #sheet = Sheet(excel_path + 'CONTROL/' + CON CMH14.xlsx')
        sheet = Sheet(excel_path + 'CONTROL/' + 'C_GO_14.xlsx')
        sheet.open_sheet_no(1)  # CON GO 14
        ftr_Time_list = sheet.get_column('E', 5)
        cmt_Time_list = sheet.get_column('H', 5)[:11]
        cmt_Text_list = sheet.get_column('I', 5)[:11]
        cmt_Time_list.pop(4)                                    # remove timestamp
        cmt_Text_list.pop(4)                                    # remove second "Glyoxal 10mM 10min"
        cmt_Text_list[4] = 'GO 10mM 10min'                      # replace to unify
        cmt_Time_list.insert(5, cmt_Time_list[5] - 480.)        # set new 8' CHEM
        cmt_Text_list.insert(5, "8' CHEM")                      # set new 8' CHEM
        cmt_Time_list.pop(7)                                    # remove timestamp
        cmt_Text_list.pop(7)                                    # remove second "SIF kurz"
        tem_Time_list = sheet.get_column('A', 7)
        tem_Value_list = sheet.get_column('B', 7)
        sheet.comments_to_csv(case_id, cmt_Time_list, cmt_Text_list)
        sheet.filter1_to_csv(case_id, ftr_Time_list)
        #sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 15, 33, 34)
        # correction of wrong temperature times
        sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 15-0, 33-14, 34-8.51)
        del sheet
        sheet = None

    elif case_id == 'C_GO_28_old':
        print('- Case:', case_id)
        # CONTROL GO 28
        # spikes
        # possibly wrong Time for temperature
        # missing "stopp"
        """ DO NOT USE THIS!
        sheet = Sheet(excel_path + 'CONTROL/' + CON CMH28.xlsx')
        sheet.open_sheet_no(1)
        ftr_Time_list = sheet.get_column('C', 5)
        cmt_Time_list = sheet.get_column('H', 5)[:9]
        cmt_Text_list = sheet.get_column('I', 5)[:9]
        print('   ', cmt_Text_list)
        cmt_Text_list[4] = 'GO 10mM 10min'                      # replace to unify
        cmt_Time_list.insert(5, cmt_Time_list[5] - 480.)
        cmt_Text_list.insert(5, "8' CHEM")
        tem_Time_list = sheet.get_column('E', 7)
        tem_Value_list = sheet.get_column('F', 7)
        sheet.comments_to_csv(case_id, cmt_Time_list, cmt_Text_list)
        sheet.filter1_to_csv(case_id, ftr_Time_list)
        sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 14, 4, 1)
        del sheet
        sheet = None
        """

    elif case_id == 'C_GO_28':
        print('- Case:', case_id)
        # CONTROL GO 28
        # spikes
        # temperature
        # missing "stopp"
        #sheet = Sheet(excel_path + 'CONTROL/' + CON C28 neu exportiert 21_7.xlsx')
        sheet = Sheet(excel_path + 'CONTROL/' + 'C_GO_28.xlsx')
        sheet.open_sheet_no(1)  # CON GO 28 neu export
        ftr_Time_list = sheet.get_column('C', 5)
        cmt_Time_list = sheet.get_column('N', 5)[:9]
        cmt_Text_list = sheet.get_column('O', 5)[:9]
        cmt_Text_list[4] = 'GO 10mM 10min'                      # replace to unify
        cmt_Time_list.insert(5, cmt_Time_list[5] - 480.)        # set new 8' CHEM
        cmt_Text_list.insert(5, "8' CHEM")                      # set new 8' CHEM
        tem_Time_list = sheet.get_column('E', 7) + sheet.get_column('I', 7) + sheet.get_column('K', 7)
        tem_Value_list = sheet.get_column('F', 7) + sheet.get_column('J', 7) + sheet.get_column('L', 7)
        # has correct temperature times
        sheet.comments_to_csv(case_id, cmt_Time_list, cmt_Text_list)
        sheet.filter1_to_csv(case_id, ftr_Time_list)
        sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 14, 4, 1)
        del sheet
        sheet = None

    elif case_id == 'C_GO_29':
        print('- Case:', case_id)
        # CONTROL GO 29
        # spikes
        # temperature
        # missing "stopp"
        #sheet = Sheet(excel_path + 'CONTROL/' + CON CMH29.xlsx')
        sheet = Sheet(excel_path + 'CONTROL/' + 'C_GO_29.xlsx')
        sheet.open_sheet_no(1)  # CON GO 29
        ftr_Time_list = sheet.get_column('C', 5)
        cmt_Time_list = sheet.get_column('H', 5)[:9]
        cmt_Text_list = sheet.get_column('I', 5)[:9]
        cmt_Text_list[4] = 'GO 10mM 10min'                      # replace to unify
        cmt_Time_list.insert(5, cmt_Time_list[5] - 480.)        # set new 8' CHEM
        cmt_Text_list.insert(5, "8' CHEM")                      # set new 8' CHEM
        tem_Time_list = sheet.get_column('E', 7)
        tem_Value_list = sheet.get_column('F', 7)
        sheet.comments_to_csv(case_id, cmt_Time_list, cmt_Text_list)
        sheet.filter1_to_csv(case_id, ftr_Time_list)
        #sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 16, 8, 55)
        del sheet
        sheet = None

    elif case_id == 'C_GO_31':
        print('- Case:', case_id)
        # CONTROL GO 31
        # spikes
        # temperature
        # missing "stopp"
        # stop with second "cold", missing "v.frey"
        #sheet = Sheet(excel_path + 'CONTROL/' + 'CON CMH31.xlsx')
        sheet = Sheet(excel_path + 'CONTROL/' + 'C_GO_31.xlsx')
        sheet.open_sheet_no(1)  # CON GO 31
        ftr_Time_list = sheet.get_column('C', 5)
        cmt_Time_list = sheet.get_column('H', 5)[:6]
        cmt_Text_list = sheet.get_column('I', 5)[:6]
        cmt_Time_list.insert(0, cmt_Time_list[0] - 0.0002)      # to unify order
        cmt_Text_list.insert(0, 'heat')                         # to unify order
        cmt_Time_list.insert(1, cmt_Time_list[1] - 0.0001)      # to unify order
        cmt_Text_list.insert(1, 'cold')                         # to unify order
        cmt_Text_list[4] = 'GO 10mM 10min'                      # replace to unify
        cmt_Time_list.insert(5, cmt_Time_list[5] - 480.)        # set new 8' CHEM
        cmt_Text_list.insert(5, "8' CHEM")                      # set new 8' CHEM
        tem_Time_list = sheet.get_column('E', 7)
        tem_Value_list = sheet.get_column('F', 7)
        sheet.comments_to_csv(case_id, cmt_Time_list, cmt_Text_list)
        sheet.filter1_to_csv(case_id, ftr_Time_list)
        #sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 19, 43, 49)
        del sheet
        sheet = None

    elif case_id == 'D_GO_20':
        print('- Case:', case_id)
        # DIABETIC GO 20
        # spikes only
        #sheet = Sheet(excel_path + 'GO/' + 'STZ GO 20 CMH - Faser 1 vom 19.07.2019 Aachen.xlsx')
        sheet = Sheet(excel_path + 'GO/' + 'D_GO_20.xlsx')
        sheet.open_sheet_no(1)
        ftr_Time_list = sheet.get_column('H', 5)
        cmt_Time_list = sheet.get_column('B', 9)
        cmt_Text_list = sheet.get_column('D', 9)
        cmt_Text_list[4] = 'GO 10mM 10min'                      # replace to unify
        cmt_Time_list.insert(5, cmt_Time_list[5] - 480.)        # set new 8' CHEM
        cmt_Text_list.insert(5, "8' CHEM")                      # set new 8' CHEM
        sheet.comments_to_csv(case_id, cmt_Time_list, cmt_Text_list)
        sheet.filter1_to_csv(case_id, ftr_Time_list)
        del sheet
        sheet = None

    elif case_id == 'D_GO_22':
        print('- Case:', case_id)
        # DIABETIC GO 22
        # spikes only
        #sheet = Sheet(excel_path + 'GO/' + 'STZ GO 22 Faser 1 vom 19_07_2019 Aachen.xlsx')
        sheet = Sheet(excel_path + 'GO/' + 'D_GO_22.xlsx')
        sheet.open_sheet_no(1)
        ftr_Time_list = sheet.get_column('U', 5)
        cmt_Time_list = sheet.get_column('P', 5)[:10]
        cmt_Text_list = sheet.get_column('Q', 5)[:10]
        cmt_Text_list[4] = 'GO 10mM 10min'                      # replace to unify
        cmt_Time_list.insert(5, cmt_Time_list[5] - 480.)        # set new 8' CHEM
        cmt_Text_list.insert(5, "8' CHEM")                      # set new 8' CHEM
        sheet.comments_to_csv(case_id, cmt_Time_list, cmt_Text_list)
        sheet.filter1_to_csv(case_id, ftr_Time_list)
        del sheet
        sheet = None

    elif case_id == 'D_MG_22':
        print('- Case:', case_id)
        # DIABETIC MG 22
        # spikes only
        #sheet = Sheet(excel_path + 'MG/' + 'STZ MG 22_20200218-Faser 2 Aachen.xlsx)
        sheet = Sheet(excel_path + 'MG/' + 'D_MG_22.xlsx')
        sheet.open_sheet_no(1)
        ftr_Time_list = sheet.get_column('T', 5)
        cmt_Time_list = sheet.get_column('O', 5)[:10]
        cmt_Text_list = sheet.get_column('P', 5)[:10]
        cmt_Text_list[3] = "8' SIF"                             # replace to unify
        cmt_Text_list[4] = 'MG 10mM 10min'                      # replace to unify
        cmt_Time_list.insert(5, cmt_Time_list[5] - 480.)        # set new 8' CHEM
        cmt_Text_list.insert(5, "8' CHEM")                      # set new 8' CHEM
        sheet.comments_to_csv(case_id, cmt_Time_list, cmt_Text_list)
        sheet.filter1_to_csv(case_id, ftr_Time_list)
        del sheet
        sheet = None

    elif case_id == 'D_MG_25':
        print('- Case:', case_id)
        # DIABETIC MG 25
        # spikes only
        #sheet = Sheet(excel_path + 'MG/' + 'STZ MG 25_20200221-Faser 2 Aachen.xlsx')
        sheet = Sheet(excel_path + 'MG/' + 'D_MG_25.xlsx')
        sheet.open_sheet_no(1)
        ftr_Time_list = sheet.get_column('U', 5)
        cmt_Time_list = sheet.get_column('P', 5)[:10]
        cmt_Text_list = sheet.get_column('Q', 5)[:10]
        cmt_Text_list[3] = "8' SIF"                             # replace to unify
        cmt_Text_list[4] = 'MG 10mM 10min'                      # replace to unify
        cmt_Time_list.insert(5, cmt_Time_list[5] - 480.)        # set new 8' CHEM
        cmt_Text_list.insert(5, "8' CHEM")                      # set new 8' CHEM
        sheet.comments_to_csv(case_id, cmt_Time_list, cmt_Text_list)
        sheet.filter1_to_csv(case_id, ftr_Time_list)
        del sheet
        sheet = None

    elif case_id == 'D_MG_08':
        print('- Case:', 'D_MG_08')
        # DIABETIC MG 08
        # spikes
        # temperature
        ###sheet = Sheet(excel_path + 'MG/' + 'STZ Fasern MG new.xlsx')
        sheet = Sheet(excel_path + 'MG/' + 'D_MG_08.xlsx')
        sheet.open_sheet_no(1)
        ftr_Time_list = sheet.get_column('C', 5)
        cmt_Time_list = sheet.get_column('H', 8)[:10]
        cmt_Text_list = sheet.get_column('I', 8)[:10]
        cmt_Text_list[3] = "8' SIF"                             # replace to unify
        cmt_Text_list[4] = 'MG 10mM 10min'                      # replace to unify
        cmt_Time_list.insert(5, cmt_Time_list[5] - 480.)        # set new 8' CHEM
        cmt_Text_list.insert(5, "8' CHEM")                      # set new 8' CHEM
        tem_Time_list = sheet.get_column('E', 7)
        tem_Value_list = sheet.get_column('F', 7)
        sheet.comments_to_csv(case_id, cmt_Time_list, cmt_Text_list)
        sheet.filter1_to_csv(case_id, ftr_Time_list)
        #sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 16, 2, 30)
        # correction of wrong temperature times
        sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 16-0, 2-15, 30-21.07)
        del sheet
        sheet = None

    elif case_id == 'D_MG_09':
        print('- Case:', 'D_MG_09')
        # DIABETIC MG 09
        # spikes
        # temperature
        ###sheet = Sheet(excel_path + 'MG/' + 'STZ Fasern MG new.xlsx')
        sheet = Sheet(excel_path + 'MG/' + 'D_MG_09.xlsx')
        #sheet.open_sheet_no(2)
        sheet.open_sheet_no(1)
        ftr_Time_list = sheet.get_column('I', 5)
        cmt_Time_list = sheet.get_column('D', 5)[:10]
        cmt_Text_list = sheet.get_column('E', 5)[:10]
        cmt_Text_list[4] = 'MG 10mM 10min'                      # replace to unify
        cmt_Time_list.insert(5, cmt_Time_list[5] - 480.)        # set new 8' CHEM
        cmt_Text_list.insert(5, "8' CHEM")                      # set new 8' CHEM
        tem_Time_list = sheet.get_column('A', 7)
        tem_Value_list = sheet.get_column('B', 7)
        sheet.comments_to_csv(case_id, cmt_Time_list, cmt_Text_list)
        sheet.filter1_to_csv(case_id, ftr_Time_list)
        #sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 10, 35, 0)
        # correction of wrong temperature times
        sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 10-0, 35-10, 0-22.15)
        del sheet
        sheet = None

    elif case_id == 'D_MG_15':
        print('- Case:', 'D_MG_15')
        # DIABETIC MG 15
        # spikes
        # temperature
        ###sheet = Sheet(excel_path + 'MG/' + 'STZ Fasern MG new.xlsx')
        sheet = Sheet(excel_path + 'MG/' + 'D_MG_15.xlsx')
        #sheet.open_sheet_no(3)
        sheet.open_sheet_no(1)
        ftr_Time_list = sheet.get_column('C', 6)
        cmt_Time_list = sheet.get_column('H', 8)[:10]
        cmt_Text_list = sheet.get_column('I', 8)[:10]
        cmt_Text_list[4] = 'MG 10mM 10min'                      # replace to unify
        cmt_Time_list.insert(5, cmt_Time_list[5] - 480.)        # set new 8' CHEM
        cmt_Text_list.insert(5, "8' CHEM")                      # set new 8' CHEM
        tem_Time_list = sheet.get_column('E', 8)
        tem_Value_list = sheet.get_column('F', 8)
        sheet.comments_to_csv(case_id, cmt_Time_list, cmt_Text_list)
        sheet.filter1_to_csv(case_id, ftr_Time_list)
        #sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 10, 7, 2)
        # correction of wrong temperature times
        sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 10-1, 7-13, 2-33.00)
        del sheet
        sheet = None

    elif case_id == 'D_MG_18':
        print('- Case:', 'D_MG_18')
        # DIABETIC MG 18
        # spikes
        # possibly wrong Time for temperature
        # contains 2 heat2 phases, second one will be deleted!
        ###sheet = Sheet(excel_path + 'MG/' + 'STZ Fasern MG new.xlsx')
        sheet = Sheet(excel_path + 'MG/' + 'D_MG_18.xlsx')
        #sheet.open_sheet_no(4)
        sheet.open_sheet_no(1)
        ftr_Time_list = sheet.get_column('C', 5)
        cmt_Time_list = sheet.get_column('H', 6)[:11]
        cmt_Text_list = sheet.get_column('I', 6)[:11]
        cmt_Text_list[4] = 'MG 10mM 10min'                      # replace to unify
        cmt_Time_list.insert(5, cmt_Time_list[5] - 480.)        # set new 8' CHEM
        cmt_Text_list.insert(5, "8' CHEM")                      # set new 8' CHEM
        cmt_Time_list.pop(9)                                    # remove timestamp
        cmt_Text_list.pop(9)                                    # remove second "heat" of heat2
        tem_Time_list = sheet.get_column('E', 7)
        tem_Value_list = sheet.get_column('F', 7)
        sheet.comments_to_csv(case_id, cmt_Time_list, cmt_Text_list)
        sheet.filter1_to_csv(case_id, ftr_Time_list)
        #sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 12, 34, 45)
        # correction of wrong temperature times
        sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 12-0, 34-47, 45-49.40)
        del sheet
        sheet = None

    elif case_id == 'D_MG_19':
        print('- Case:', 'D_MG_19')
        # DIABETIC MG 19
        # spikes
        # temperature
        ###sheet = Sheet(excel_path + 'MG/' + 'STZ Fasern MG new.xlsx')
        sheet = Sheet(excel_path + 'MG/' + 'D_MG_19.xlsx')
        #sheet.open_sheet_no(5)
        sheet.open_sheet_no(1)
        ftr_Time_list = sheet.get_column('I', 6)
        cmt_Time_list = sheet.get_column('D', 10)[:10]
        cmt_Text_list = sheet.get_column('E', 10)[:10]
        cmt_Text_list[3] = "8' SIF"                             # replace to unify
        cmt_Text_list[4] = 'MG 10mM 10min'                      # replace to unify
        cmt_Time_list.insert(5, cmt_Time_list[5] - 480.)        # set new 8' CHEM
        cmt_Text_list.insert(5, "8' CHEM")                      # set new 8' CHEM
        tem_Time_list = sheet.get_column('A', 8)
        tem_Value_list = sheet.get_column('B', 8)
        sheet.comments_to_csv(case_id, cmt_Time_list, cmt_Text_list)
        sheet.filter1_to_csv(case_id, ftr_Time_list)
        #sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 13, 19, 2)
        # correction of wrong temperature times
        sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 13-1, 19-1, 2-47.18)
        del sheet
        sheet = None

    elif case_id == 'D_MG_24':
        print('- Case:', 'D_MG_24')
        # DIABETIC MG 24
        # contains 'auf170'
        ###sheet = Sheet(excel_path + 'MG/' + 'STZ Fasern MG new.xlsx')
        sheet = Sheet(excel_path + 'MG/' + 'D_MG_24.xlsx')
        #sheet.open_sheet_no(6)
        sheet.open_sheet_no(1)
        ftr_Time_list = sheet.get_column('C', 5)
        cmt_Time_list = sheet.get_column('H', 8)[:11]
        cmt_Text_list = sheet.get_column('I', 8)[:11]
        cmt_Time_list.pop(2)                                    # remove timestamp
        cmt_Text_list.pop(2)                                    # remove second "auf170"
        cmt_Text_list[3] = "8' SIF"                             # replace to unify
        cmt_Text_list[4] = 'MG 10mM 10min'                      # replace to unify
        cmt_Time_list.insert(5, cmt_Time_list[5] - 480.)        # set new 8' CHEM
        cmt_Text_list.insert(5, "8' CHEM")                      # set new 8' CHEM
        tem_Time_list = sheet.get_column('E', 7)
        tem_Value_list = sheet.get_column('F', 7)
        sheet.comments_to_csv(case_id, cmt_Time_list, cmt_Text_list)
        sheet.filter1_to_csv(case_id, ftr_Time_list)
        #sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 10, 22, 43)
        # correction of wrong temperature times
        sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 10-0, 22-9, 43-3.98)
        del sheet
        sheet = None

    elif case_id == 'D_MG_28':
        print('- Case:', 'D_MG_28')
        # DIABETIC MG 28
        # spikes
        # temperature
        # contains 'auf170'
        # missing "8' SIF"
        ###sheet = Sheet(excel_path + 'MG/' + 'STZ Fasern MG new.xlsx')
        sheet = Sheet(excel_path + 'MG/' + 'D_MG_28.xlsx')
        #sheet.open_sheet_no(7)
        sheet.open_sheet_no(1)
        ftr_Time_list = sheet.get_column('C', 5)
        cmt_Time_list = sheet.get_column('H', 8)[:10]
        cmt_Text_list = sheet.get_column('I', 8)[:10]
        cmt_Time_list.pop(2)                                    # remove timestamp
        cmt_Text_list.pop(2)                                    # remove second "auf170"
        cmt_Time_list.insert(3, cmt_Time_list[3] - 480.)        # to unify order
        cmt_Text_list.insert(3, "8' SIF")                       # to unify order
        cmt_Text_list[4] = 'MG 10mM 10min'                      # replace to unify
        cmt_Time_list.insert(5, cmt_Time_list[5] - 480.)        # set new 8' CHEM
        cmt_Text_list.insert(5, "8' CHEM")                      # set new 8' CHEM
        tem_Time_list = sheet.get_column('E', 7)
        tem_Value_list = sheet.get_column('F', 7)
        sheet.comments_to_csv(case_id, cmt_Time_list, cmt_Text_list)
        sheet.filter1_to_csv(case_id, ftr_Time_list)
        #sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 13, 10, 3)
        # correction of wrong temperature times
        sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 13-0, 10-12, 3-2.82)
        del sheet
        sheet = None

    elif case_id == 'D_MG_36':
        print('- Case:', 'D_MG_36')
        # DIABETIC MG 36
        # spikes
        # temperature
        ###sheet = Sheet(excel_path + 'MG/' + 'STZ Fasern MG new.xlsx')
        sheet = Sheet(excel_path + 'MG/' + 'D_MG_36.xlsx')
        #sheet.open_sheet_no(8)
        sheet.open_sheet_no(1)
        ftr_Time_list = sheet.get_column('C', 6)
        cmt_Time_list = sheet.get_column('H', 6)[:10]
        cmt_Text_list = sheet.get_column('I', 6)[:10]
        cmt_Text_list[3] = "8' SIF"                             # replace to unify
        cmt_Text_list[4] = 'MG 10mM 10min'                      # replace to unify
        cmt_Time_list.insert(5, cmt_Time_list[5] - 480.)        # set new 8' CHEM
        cmt_Text_list.insert(5, "8' CHEM")                      # set new 8' CHEM
        tem_Time_list = sheet.get_column('E', 8)
        tem_Value_list = sheet.get_column('F', 8)
        sheet.comments_to_csv(case_id, cmt_Time_list, cmt_Text_list)
        sheet.filter1_to_csv(case_id, ftr_Time_list)
        #sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 17, 18, 52)
        # correction of wrong temperature times
        sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 17-0, 18-7, 52-33.30)
        del sheet
        sheet = None

    elif case_id == 'D_MG_39':
        print('- Case:', 'D_MG_39')
        # DIABETIC MG 39
        # spikes
        # temperature
        ###sheet = Sheet(excel_path + 'MG/' + 'STZ Fasern MG new.xlsx')
        sheet = Sheet(excel_path + 'MG/' + 'D_MG_several.xlsx')
        sheet.open_sheet_no(9)
        ftr_Time_list = sheet.get_column('M', 5)
        cmt_Time_list = sheet.get_column('H', 7)[:10]
        cmt_Text_list = sheet.get_column('I', 7)[:10]
        cmt_Text_list[3] = "8' SIF"                             # replace to unify
        cmt_Text_list[4] = 'MG 10mM 10min'                      # replace to unify
        cmt_Time_list.insert(5, cmt_Time_list[5] - 480.)        # set new 8' CHEM
        cmt_Text_list.insert(5, "8' CHEM")                      # set new 8' CHEM
        tem_Time_list = sheet.get_column('E', 7)
        tem_Value_list = sheet.get_column('F', 7)
        sheet.comments_to_csv(case_id, cmt_Time_list, cmt_Text_list)
        sheet.filter1_to_csv(case_id, ftr_Time_list)
        #sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 19, 32, 11)
        # correction of wrong temperature times
        sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 19-0, 32-17, 11-0.00)
        del sheet
        sheet = None

    elif case_id == 'C_Adelta_21':
        print('- Case:', case_id)
        # CONTROL A delta Fiber 21
        # spikes
        # missing "stopp"
        # missing second cold
        ###sheet = Sheet(excel_path + 'CONTROL/' + 'CON Adelta A21.xlsx')
        sheet = Sheet(excel_path + 'CONTROL/' + 'C_Adelta_21.xlsx')
        sheet.open_sheet_no(1)
        ftr_Time_list = sheet.get_column('C', 5)
        cmt_Time_list = sheet.get_column('L', 5)
        cmt_Text_list = sheet.get_column('M', 5)
        cmt_Time_list.insert(3, cmt_Time_list[3] - 480.)        # to unify order
        cmt_Text_list.insert(3, "8' SIF")                       # to unify order
        cmt_Text_list[4] = 'MG 10mM 10min'                      # replace to unify
        cmt_Time_list.insert(5, cmt_Time_list[5] - 480.)        # set new 8' CHEM
        cmt_Text_list.insert(5, "8' CHEM")                      # set new 8' CHEM
        tem_Time_list = sheet.get_column('E', 7) + sheet.get_column('G', 7) + sheet.get_column('I', 7)
        tem_Value_list = sheet.get_column('F', 7) + sheet.get_column('H', 7) + sheet.get_column('J', 7)
        sheet.comments_to_csv(case_id, cmt_Time_list, cmt_Text_list)
        sheet.filter1_to_csv(case_id, ftr_Time_list)
        sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 13, 35, 49)
        del sheet
        sheet = None

    elif case_id == 'C_Adelta_24':
        print('- Case:', case_id)
        # CONTROL A delta Fiber 24
        # spikes
        # missing "stopp"
        ###sheet = Sheet(excel_path + 'CONTROL/' + 'CON Adelta A24.xlsx')
        sheet = Sheet(excel_path + 'CONTROL/' + 'C_Adelta_24.xlsx')
        sheet.open_sheet_no(1)
        ftr_Time_list = sheet.get_column('C', 5)
        cmt_Time_list = sheet.get_column('N', 5)
        cmt_Text_list = sheet.get_column('O', 5)
        cmt_Time_list.insert(3, cmt_Time_list[3] - 480.)        # to unify order
        cmt_Text_list.insert(3, "8' SIF")                       # to unify order
        cmt_Text_list[4] = 'MG 10mM 10min'                      # replace to unify
        cmt_Time_list.insert(5, cmt_Time_list[5] - 480.)        # set new 8' CHEM
        cmt_Text_list.insert(5, "8' CHEM")                      # set new 8' CHEM
        tem_Time_list = sheet.get_column('E', 7) + sheet.get_column('G', 7) + sheet.get_column('I', 7)
        tem_Value_list = sheet.get_column('F', 7) + sheet.get_column('H', 7) + sheet.get_column('J', 7)
        sheet.comments_to_csv(case_id, cmt_Time_list, cmt_Text_list)
        sheet.filter1_to_csv(case_id, ftr_Time_list)
        sheet.temperature_to_csv(case_id, tem_Time_list, tem_Value_list, 12, 1, 26)
        del sheet
        sheet = None

    else:
        print('No such case as', case_id)
        return
    '''

    del ftr_Time_list, cmt_Time_list, cmt_Text_list, tem_Time_list, tem_Value_list
    if sheet is not None:
        del sheet

    # more csv-files for relevant sections
    file_spikes = csv_path + case_id + '_filter1.csv'
    file_comments = csv_path + case_id + '_comments.csv'
    file_temperature = csv_path + case_id + '_temperature.csv'  # optional
    file_spikes_SIF = csv_path + case_id + '_spikes_SIF.csv'
    file_spikes_CHEM = csv_path + case_id + '_spikes_CHEM.csv'
    file_isi_SIF = csv_path + case_id + '_isi_SIF.csv'
    file_isi_CHEM = csv_path + case_id + '_isi_CHEM.csv'

    try:
        # All spikes
        spikes = read_csv_columns(file_spikes, ['Time'])['Time']
        del file_spikes

        # Relevant spikes
        comments = read_csv_columns(file_comments, ['Time', 'Text'])
        del file_comments
        comments_time = comments['Time']
        comments_text = comments['Text']
        del comments
        SIF_start = comments_time[3]
        SIF_end = SIF_start + relevant_size
        CHEM_start = comments_time[5]
        CHEM_end = CHEM_start + relevant_size
        spikes_SIF = list(filter(lambda x: SIF_start <= x and x < SIF_end, spikes))
        spikes_CHEM = list(filter(lambda x: CHEM_start <= x and x < CHEM_end, spikes))
        save_to_file(file_spikes_SIF, spikes_SIF)
        save_to_file(file_spikes_CHEM, spikes_CHEM)
        del SIF_end
        del CHEM_end
        del file_spikes_SIF
        del file_spikes_CHEM

        # The ISI would only include the time between the first spike and the last spike.
        # Therefor we would lose all time between "start and first spike" as well as between
        # "last spike and end". To compensate for that, we will add the last spike of the previous
        # section to the spikes list, when we calculate the ISI.
        spikes_pre_SIF = list(filter(lambda x: x < SIF_start, spikes))
        spikes_pre_CHEM = list(filter(lambda x: x < CHEM_start, spikes))
        isi_SIF = np.diff([spikes_pre_SIF[-1]] + spikes_SIF)
        isi_CHEM = np.diff([spikes_pre_CHEM[-1]] + spikes_CHEM)
        save_to_file(file_isi_SIF, isi_SIF)
        save_to_file(file_isi_CHEM, isi_CHEM)
        #print('SIF: ', round(SIF_start-spikes_pre_SIF[-1], 2), '; SUM:', round(sum(isi_SIF), 2))
        #print('CHEM:', round(CHEM_start-spikes_pre_CHEM[-1], 2), '; SUM:', round(sum(isi_CHEM), 2))
        del SIF_start
        del CHEM_start
        del spikes_SIF
        del spikes_CHEM
        del spikes_pre_SIF
        del spikes_pre_CHEM
        del isi_SIF
        del isi_CHEM

        # temperature (optional!)
        try:
            csv_heat = read_csv_columns(file_temperature, ['Time', 'Value'])
            # continues only if not FileNotFoundError

            heat1 = 0                   # fixed index of the first heat in sanitized comments
            # heat1 = comments_text.index('heat')
            heat1_start = comments_time[heat1]
            heat1_end = comments_time[heat1+1]
            spikes_heat1 = list(filter(lambda x: heat1_start <= x and x < heat1_end, spikes))
            time_heat1 = list(filter(lambda x: heat1_start <= x and x < heat1_end, csv_heat['Time']))
            #save_to_file(csv_path + case_id + '_spikes_heat1.csv', spikes_heat1)
            del heat1_start, heat1_end
            del spikes_heat1

            if len(time_heat1) > 1:
                value_heat1 = csv_heat['Value'][csv_heat['Time'].index(time_heat1[0]) :
                                                    csv_heat['Time'].index(time_heat1[-1])]
                #with open(csv_path + case_id + '_temperature_heat1.csv', mode='wt') as fw:
                #    fw.write('Time;Value' + '\n')
                #    for (time, value) in zip(heat1_tem_time, heat1_tem_value):
                #        fw.write('%.4f;%.2f\n' % (time, value))
                del value_heat1
            del time_heat1

            if comments_text.count('heat') > 1:
                # searching second heat
                heat2 = comments_text.index('heat', heat1+2)
                del heat1
                heat2_start = comments_time[heat2]
                if (heat2 < len(comments_text)-1):   # at least 1 entry after it
                    heat2_end = comments_time[heat2+1]
                    spikes_heat2 = list(filter(lambda x: heat2_start <= x and x < heat2_end, spikes))
                    time_heat2 = list(filter(lambda x: heat2_start <= x and x < heat2_end, csv_heat['Time']))
                    del heat2_end
                else:
                    spikes_heat2 = list(filter(lambda x: heat2_start <= x, spikes))
                    time_heat2 = list(filter(lambda x: heat2_start <= x, csv_heat['Time']))
                #save_to_file(csv_path + case_id + '_spikes_heat2.csv', spikes_heat2)
                del heat2_start
                del spikes_heat2

                if len(time_heat2) > 1:
                    value_heat2 = csv_heat['Value'][csv_heat['Time'].index(time_heat2[0]) :
                                                    csv_heat['Time'].index(time_heat2[-1])]
                    #with open(csv_path + case_id + '_temperature_heat2.csv', mode='wt') as fw:
                    #    fw.write('Time;Value' + '\n')
                    #    for (time, value) in zip(heat2_tem_time, heat2_tem_value):
                    #        fw.write('%.4f;%.2f\n' % (time, value))
                    del value_heat2
                del time_heat2
            del csv_heat
        except FileNotFoundError:
            'There is no temperature.csv file and that is okay!'    # do nothing
        except Exception as e:
            print(type(e))
            print(e)
        finally:
            del spikes
            del comments_time
            del comments_text

    except Exception as e:
        print(type(e))
        print(e)

